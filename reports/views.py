import csv
import io
from datetime import date

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from core.decorators import require_role_decorator
from reports.services.reports_service import RoomReportService, EquipmentReportService


def _get_date_range(request):
    today = date.today()
    default_from = today.replace(day=1)
    default_to   = today

    try:
        date_from = date.fromisoformat(request.GET.get('date_from', ''))
    except (ValueError, TypeError):
        date_from = default_from

    try:
        date_to = date.fromisoformat(request.GET.get('date_to', ''))
    except (ValueError, TypeError):
        date_to = default_to

    if date_from > date_to:
        date_from, date_to = date_to, date_from

    return date_from, date_to


@login_required(login_url='login')
@require_role_decorator(roles=['operator', 'approver'])
def reports_page(request):
    date_from, date_to = _get_date_range(request)
    return render(request, 'reports/reports.html', {
        'date_from': date_from.isoformat(),
        'date_to':   date_to.isoformat(),
    })


@login_required(login_url='login')
@require_role_decorator(roles=['operator', 'approver'])
def rooms_report_data(request):
    date_from, date_to = _get_date_range(request)
    data = RoomReportService.get_report(date_from, date_to)
    return JsonResponse(data)


@login_required(login_url='login')
@require_role_decorator(roles=['operator', 'approver'])
def equipment_report_data(request):
    date_from, date_to = _get_date_range(request)
    data = EquipmentReportService.get_report(date_from, date_to)
    return JsonResponse(data)


@login_required(login_url='login')
@require_role_decorator(roles=['operator', 'approver'])
def export_csv(request):
    date_from, date_to = _get_date_range(request)
    report_type = request.GET.get('type', 'rooms')

    if report_type == 'rooms':
        data   = RoomReportService.get_report(date_from, date_to)
        fields = ['name', 'building', 'type', 'bookings_count', 'total_hours', 'load_pct', 'peak_day']
        headers = ['Аудитория', 'Корпус', 'Тип', 'Мероприятий', 'Часов', 'Загруженность %', 'Пиковый день']
        filename = f'rooms_report_{date_from}_{date_to}.csv'
    else:
        data   = EquipmentReportService.get_report(date_from, date_to)
        fields = ['name', 'model', 'inventory_number', 'type', 'bookings_count', 'total_hours', 'load_pct']
        headers = ['Наименование', 'Модель', 'Инв. номер', 'Тип', 'Использований', 'Часов', 'Востребованность %']
        filename = f'equipment_report_{date_from}_{date_to}.csv'

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'

    writer = csv.writer(response)
    writer.writerow(headers)
    for item in data.get('items', []):
        writer.writerow([item.get(f, '') for f in fields])

    return response


@login_required(login_url='login')
@require_role_decorator(roles=['operator', 'approver'])
def export_xlsx(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        return HttpResponse('openpyxl не установлен. pip install openpyxl', status=500)

    date_from, date_to = _get_date_range(request)
    report_type = request.GET.get('type', 'rooms')

    if report_type == 'rooms':
        data    = RoomReportService.get_report(date_from, date_to)
        fields  = ['name', 'building', 'type', 'bookings_count', 'total_hours', 'load_pct', 'peak_day']
        headers = ['Аудитория', 'Корпус', 'Тип', 'Мероприятий', 'Часов', 'Загруженность %', 'Пиковый день']
        sheet_title = 'Аудитории'
        filename = f'rooms_report_{date_from}_{date_to}.xlsx'
    else:
        data    = EquipmentReportService.get_report(date_from, date_to)
        fields  = ['name', 'model', 'inventory_number', 'type', 'bookings_count', 'total_hours', 'load_pct']
        headers = ['Наименование', 'Модель', 'Инв. номер', 'Тип', 'Использований', 'Часов', 'Востребованность %']
        sheet_title = 'Оборудование'
        filename = f'equipment_report_{date_from}_{date_to}.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws.merge_cells('A1:G1')
    ws['A1'] = f'Отчёт: {sheet_title} · {date_from} — {date_to}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='left')

    header_fill = PatternFill('solid', fgColor='1E4BA3')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, item in enumerate(data.get('items', []), 3):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(field, ''))

    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value and not isinstance(cell, MergedCell):
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_len + 4, 40)
        if adjusted_width > 0 and hasattr(cell, 'column_letter'):
            ws.column_dimensions[cell.column_letter].width = adjusted_width
        else:
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response